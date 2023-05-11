import json
import requests
from openpyxl import load_workbook
import csv

def write_in_csv(json_data: json):
    fileNameToSave = './data.csv'
    key_data = ['operationToken', 'firstName', 'lastName', 'middleName', 'telephone', 'car', 'passport', 'dayBirth', 'monthBirth', 'yearBirth', 'snils', 'inn', 'base', 'id', 'information']
    operationToken = json_data['operationToken']
    list_of_dicts = []

    for i in json_data['resultList']:
        dict = {'operationToken': operationToken}
        for k,v in i.items():
            dict[k] = v
        list_of_dicts.append(dict)

    with open(fileNameToSave, '+a', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, delimiter = ',', fieldnames = key_data)
            # print('count: ' + str(len(list_of_dicts)))
            for element in list_of_dicts:
                writer.writerow(element)

def response_solari_api_by_phone(phone: str)-> json:
    url = "https://api.id-x.org/idx/api2/special/check"
    payload = json.dumps({
    "accessKey": "ключ",
    "secretKey": "ключ",
    "telephone": phone
    })
    headers = {
    'Content-Type': 'application/json'
    }
    response = requests.request("POST", url, headers=headers, data=payload)
    return response


def get_phone_numbers(file: str)-> list:
    wb = load_workbook(file)
    sheet = wb['Sheet1']
    sheet.title
    return [str(sheet.cell(row=i, column=1).value)[1:] for i in range(23,24)]


phone_numbers = get_phone_numbers('/Users/knapptan/Desktop/IDX_work/sravnenie_servisov/sravnenie_v3/base_v3.xlsx')
count = 22
puss = '/Users/knapptan/Desktop/IDX_work/sravnenie_servisov/sravnenie_v3/json_logs/'


for phone_number in phone_numbers:
    file_name = 'data_salari_num'+str(count)+'.json'
    print(phone_number)
    response = response_solari_api_by_phone(phone_number)
    write_in_csv(response.json())
    with open(puss + file_name, '+w') as file:
        json.dump(json.loads(response.text), file,ensure_ascii=False, indent=4)
    count += 1
    # response = response_solari_api_by_phone(phone_number)

with open(puss + file_name, 'r' ) as f:
        data = f.read()
        json_data = json.loads(data)

key_list_GG = []
for element in json_data.get('data'):
      key_list_GG.append(element)

print(key_list_GG)

# writekey_list_GG_csv(json_data['data'], 'data_getscam.csv')
# save('data_getscam.csv', json_data['data'])
